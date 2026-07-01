"""
Phase 2 - Duplicate Detection (exact duplicates only)

Reads the inventory database built by scan_inventory.py, groups files that
are byte-for-byte identical, and exports a review spreadsheet with a
suggested "keep" copy for each duplicate group.

This does NOT delete anything. It only flags. Deletion is a manual step
you take after reviewing duplicates.xlsx (a delete_confirmed.py helper can
be added later once you're comfortable with the workflow).

Usage:
    python find_duplicates.py
"""
import sqlite3
from pathlib import Path

import config
from scan_inventory import full_hash  # reuse the same hashing function


def ensure_full_hashes(conn):
    """
    Exact-duplicate detection requires a full_hash for every file that
    shares a (size, partial_hash) with at least one other file. scan_inventory.py
    only fills this in opportunistically during a single scan run, so files
    matched across separate runs may still be missing it. Fill any gaps now.
    """
    cur = conn.cursor()
    cur.execute(
        """SELECT size_bytes, partial_hash FROM files
           WHERE partial_hash IS NOT NULL
           GROUP BY size_bytes, partial_hash HAVING COUNT(*) > 1"""
    )
    candidate_groups = cur.fetchall()

    filled = 0
    for size_bytes, p_hash in candidate_groups:
        cur.execute(
            "SELECT id, path, full_hash FROM files WHERE size_bytes=? AND partial_hash=?",
            (size_bytes, p_hash),
        )
        rows = cur.fetchall()
        for file_id, path, existing_hash in rows:
            if existing_hash:
                continue
            h = full_hash(Path(path))
            if h:
                cur.execute("UPDATE files SET full_hash=? WHERE id=?", (h, file_id))
                filled += 1
    conn.commit()
    if filled:
        print(f"Computed {filled} additional full hashes for cross-run matches.")


def suggest_keep(paths_rows):
    """
    Given rows (path, created, category, size_bytes) for one duplicate group,
    return the index of the row to suggest keeping.
    Priority: shortest path -> oldest created date.
    """
    def sort_key(row):
        path, created, category, size_bytes = row
        return (len(path), created or "9999")

    ranked = sorted(range(len(paths_rows)), key=lambda i: sort_key(paths_rows[i]))
    return ranked[0]


def find_duplicates():
    conn = sqlite3.connect(config.DB_PATH)
    ensure_full_hashes(conn)

    cur = conn.cursor()
    cur.execute(
        """SELECT full_hash FROM files
           WHERE full_hash IS NOT NULL
           GROUP BY full_hash HAVING COUNT(*) > 1"""
    )
    dup_hashes = [r[0] for r in cur.fetchall()]

    groups = []
    total_reclaimable = 0

    for group_id, h in enumerate(dup_hashes, start=1):
        cur.execute(
            "SELECT path, created, category, size_bytes FROM files WHERE full_hash=?",
            (h,),
        )
        rows = cur.fetchall()
        keep_idx = suggest_keep(rows)
        size = rows[0][3]
        reclaimable = size * (len(rows) - 1)
        total_reclaimable += reclaimable

        for i, (path, created, category, size_bytes) in enumerate(rows):
            groups.append({
                "group_id": group_id,
                "path": path,
                "category": category,
                "size_MB": round(size_bytes / (1024 * 1024), 2),
                "created": created,
                "suggested_action": "KEEP" if i == keep_idx else "DELETE",
            })

    conn.close()
    return groups, total_reclaimable


def export_report(groups, total_reclaimable):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill

    out_path = Path(config.OUTPUT_DIR)
    out_path.mkdir(exist_ok=True)
    out_file = out_path / "duplicates.xlsx"

    if not groups:
        df = pd.DataFrame(columns=["group_id", "path", "category", "size_MB", "created", "suggested_action"])
    else:
        df = pd.DataFrame(groups)

    df.to_excel(out_file, index=False, sheet_name="Duplicates")

    wb = load_workbook(out_file)
    ws = wb["Duplicates"]
    red_fill = PatternFill("solid", start_color="FFF2CC", end_color="FFF2CC")
    green_fill = PatternFill("solid", start_color="D9EAD3", end_color="D9EAD3")

    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)

    action_col = None
    for i, cell in enumerate(ws[1], start=1):
        if cell.value == "suggested_action":
            action_col = i
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")
        if action_col:
            action_cell = row[action_col - 1]
            action_cell.fill = green_fill if action_cell.value == "KEEP" else red_fill

    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
    wb.save(out_file)

    num_groups = df["group_id"].nunique() if not df.empty else 0
    print(f"\nDuplicate groups found : {num_groups}")
    print(f"Total duplicate files  : {len(df[df['suggested_action']=='DELETE']) if not df.empty else 0}")
    print(f"Reclaimable space      : {total_reclaimable / (1024**3):.2f} GB")
    print(f"Report exported to     : {out_file}")


if __name__ == "__main__":
    print("Scanning database for exact duplicates...")
    groups, total_reclaimable = find_duplicates()
    export_report(groups, total_reclaimable)
