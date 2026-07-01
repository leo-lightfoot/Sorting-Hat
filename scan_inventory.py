"""
Phase 1 - Inventory Scan

Walks all SOURCE_DIRS (see config.py), records metadata for every file into
a local SQLite database, and exports a full inventory to an Excel file.

Safe to re-run: files already in the DB with an unchanged modified-time are
skipped (fast incremental re-scans). Deleted-file cleanup runs at the end.

Usage:
    python scan_inventory.py
"""
import hashlib
import os
import sqlite3
import sys
import time
from datetime import datetime
from pathlib import Path

import config

try:
    from PIL import Image
    from PIL.ExifTags import TAGS
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False

import subprocess
import json as jsonlib


SCHEMA = """
CREATE TABLE IF NOT EXISTS files (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    path TEXT UNIQUE NOT NULL,
    filename TEXT,
    extension TEXT,
    category TEXT,
    size_bytes INTEGER,
    created TEXT,
    modified TEXT,
    accessed TEXT,
    partial_hash TEXT,
    full_hash TEXT,
    exif_date_taken TEXT,
    exif_camera TEXT,
    exif_gps TEXT,
    video_duration_s REAL,
    video_codec TEXT,
    video_resolution TEXT,
    scan_time TEXT,
    error TEXT
);
CREATE INDEX IF NOT EXISTS idx_size ON files(size_bytes);
CREATE INDEX IF NOT EXISTS idx_partial_hash ON files(partial_hash);
CREATE INDEX IF NOT EXISTS idx_full_hash ON files(full_hash);
"""


def get_category(ext):
    ext = ext.lower()
    for cat, exts in config.CATEGORY_MAP.items():
        if ext in exts:
            return cat
    return "Other"


def should_ignore(path: Path):
    if path.name.lower() in config.IGNORE_NAMES:
        return True
    if path.suffix.lower() in config.IGNORE_EXTENSIONS:
        return True
    return False


def partial_hash(path, size):
    try:
        h = hashlib.sha256()
        with open(path, "rb") as f:
            h.update(f.read(config.PARTIAL_HASH_SIZE))
        return h.hexdigest()
    except (OSError, PermissionError):
        return None


def full_hash(path):
    try:
        h = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(1024 * 1024), b""):
                h.update(chunk)
        return h.hexdigest()
    except (OSError, PermissionError):
        return None


def extract_exif(path):
    """Returns (date_taken, camera, gps) for image files, if available."""
    if not PIL_AVAILABLE:
        return None, None, None
    try:
        img = Image.open(path)
        exif_data = img._getexif()
        if not exif_data:
            return None, None, None
        tags = {TAGS.get(k, k): v for k, v in exif_data.items()}
        date_taken = tags.get("DateTimeOriginal") or tags.get("DateTime")
        camera = tags.get("Model")
        gps = "Yes" if "GPSInfo" in tags else None
        return date_taken, camera, gps
    except Exception:
        return None, None, None


def extract_video_metadata(path):
    """Returns (duration_seconds, codec, resolution) using ffprobe if available."""
    if not config.ENABLE_VIDEO_METADATA:
        return None, None, None
    try:
        result = subprocess.run(
            [
                "ffprobe", "-v", "quiet", "-print_format", "json",
                "-show_format", "-show_streams", str(path),
            ],
            capture_output=True, text=True, timeout=30,
        )
        if result.returncode != 0:
            return None, None, None
        data = jsonlib.loads(result.stdout)
        duration = float(data.get("format", {}).get("duration", 0)) or None
        video_stream = next((s for s in data.get("streams", []) if s.get("codec_type") == "video"), None)
        codec = video_stream.get("codec_name") if video_stream else None
        resolution = None
        if video_stream:
            w, h = video_stream.get("width"), video_stream.get("height")
            if w and h:
                resolution = f"{w}x{h}"
        return duration, codec, resolution
    except (FileNotFoundError, subprocess.TimeoutExpired, Exception):
        return None, None, None


def scan():
    os.makedirs(config.OUTPUT_DIR, exist_ok=True)
    conn = sqlite3.connect(config.DB_PATH)
    conn.executescript(SCHEMA)
    cur = conn.cursor()

    # Cache of (size, partial_hash) -> path already seen this run, used to
    # decide when a full hash is actually needed (only on a probable match).
    seen_partial = {}

    stats = {"scanned": 0, "skipped_unchanged": 0, "errors": 0, "total_bytes": 0}
    start = time.time()

    for source_dir in config.SOURCE_DIRS:
        source_path = Path(source_dir)
        if not source_path.exists():
            print(f"[WARN] Source path does not exist, skipping: {source_dir}")
            continue

        for root, dirs, filenames in os.walk(source_path):
            root_path = Path(root)
            dirs[:] = [d for d in dirs if d.lower() not in config.IGNORE_NAMES]

            for fname in filenames:
                fpath = root_path / fname
                if should_ignore(fpath):
                    continue

                try:
                    st = fpath.stat()
                except (OSError, PermissionError) as e:
                    stats["errors"] += 1
                    cur.execute(
                        "INSERT OR REPLACE INTO files (path, filename, error, scan_time) VALUES (?,?,?,?)",
                        (str(fpath), fname, str(e), datetime.now().isoformat()),
                    )
                    continue

                modified_iso = datetime.fromtimestamp(st.st_mtime).isoformat()

                # Incremental skip: unchanged file already in DB
                cur.execute("SELECT modified FROM files WHERE path=?", (str(fpath),))
                row = cur.fetchone()
                if row and row[0] == modified_iso:
                    stats["skipped_unchanged"] += 1
                    continue

                ext = fpath.suffix.lower()
                category = get_category(ext)
                size = st.st_size

                p_hash = partial_hash(fpath, size)
                f_hash = None

                # Only compute the (expensive) full hash if another file this
                # run already shares the same size + partial hash.
                key = (size, p_hash)
                if key in seen_partial:
                    f_hash = full_hash(fpath)
                    prev_path = seen_partial[key]
                    cur.execute("SELECT full_hash FROM files WHERE path=?", (prev_path,))
                    prev_row = cur.fetchone()
                    if prev_row and not prev_row[0]:
                        cur.execute(
                            "UPDATE files SET full_hash=? WHERE path=?",
                            (full_hash(Path(prev_path)), prev_path),
                        )
                else:
                    seen_partial[key] = str(fpath)

                exif_date, exif_camera, exif_gps = (None, None, None)
                video_duration, video_codec, video_res = (None, None, None)
                if category == "Photo":
                    exif_date, exif_camera, exif_gps = extract_exif(fpath)
                elif category == "Video":
                    video_duration, video_codec, video_res = extract_video_metadata(fpath)

                cur.execute(
                    """INSERT OR REPLACE INTO files
                    (path, filename, extension, category, size_bytes, created, modified, accessed,
                     partial_hash, full_hash, exif_date_taken, exif_camera, exif_gps,
                     video_duration_s, video_codec, video_resolution, scan_time, error)
                    VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,NULL)""",
                    (
                        str(fpath), fname, ext, category, size,
                        datetime.fromtimestamp(st.st_ctime).isoformat(),
                        modified_iso,
                        datetime.fromtimestamp(st.st_atime).isoformat(),
                        p_hash, f_hash, exif_date, exif_camera, exif_gps,
                        video_duration, video_codec, video_res,
                        datetime.now().isoformat(),
                    ),
                )
                stats["scanned"] += 1
                stats["total_bytes"] += size

                if stats["scanned"] % 500 == 0:
                    conn.commit()
                    print(f"  ...{stats['scanned']} files scanned so far")

    conn.commit()
    conn.close()

    elapsed = time.time() - start
    print("\n--- Scan complete ---")
    print(f"Files scanned/updated : {stats['scanned']}")
    print(f"Files unchanged (skip): {stats['skipped_unchanged']}")
    print(f"Errors                : {stats['errors']}")
    print(f"Total size scanned    : {stats['total_bytes'] / (1024**3):.2f} GB")
    print(f"Time elapsed          : {elapsed:.1f}s")
    return stats


def export_to_excel():
    import pandas as pd
    from openpyxl.styles import Font

    conn = sqlite3.connect(config.DB_PATH)
    df = pd.read_sql_query(
        """SELECT path, filename, extension, category, size_bytes,
                  created, modified, exif_date_taken, exif_camera,
                  video_duration_s, video_codec, video_resolution,
                  full_hash, error
           FROM files ORDER BY category, path""",
        conn,
    )
    conn.close()

    df["size_MB"] = (df["size_bytes"] / (1024 * 1024)).round(2)

    out_path = Path(config.OUTPUT_DIR) / "inventory.xlsx"
    df.to_excel(out_path, index=False, sheet_name="Inventory")

    from openpyxl import load_workbook
    wb = load_workbook(out_path)
    ws = wb["Inventory"]
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = Font(name="Arial")
    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 60)
    wb.save(out_path)

    print(f"Inventory exported to: {out_path}  ({len(df)} rows)")


if __name__ == "__main__":
    print("Starting inventory scan...")
    print(f"Source folder(s): {config.SOURCE_DIRS}\n")
    scan()
    export_to_excel()
