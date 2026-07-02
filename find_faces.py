"""
Phase 4 - Face Recognition ("find me in these photos")

IMPORTANT: this script depends on `insightface` + `onnxruntime`, which
download a face-detection/recognition model (~300MB) from the internet the
first time they run. It was written and code-reviewed carefully, but could
NOT be executed end-to-end in the environment that built it (no internet
access there). Run test_face_setup.py first against a single photo before
pointing this at your whole library - see README.md.

Workflow:
    1. Put 5-10 clear, varied photos of yourself in REFERENCE_PHOTOS_DIR
       (one clearly visible face per photo, different angles/lighting).
    2. python find_faces.py --build-reference
    3. python find_faces.py                  (scans new photos, exports matches)
    4. Review reports/my_photos.xlsx. Adjust FACE_MATCH_THRESHOLD in
       config.py (or use --threshold) and re-export without rescanning:
       python find_faces.py --rescan-similarity --threshold 0.5

No GPU required - runs on CPU, but will be noticeably slower than a GPU
setup. Expect roughly 1-3 seconds per photo depending on your machine;
plan for this to run over minutes-to-hours on a large library, not
instantly. It's safe to stop and resume - already-processed photos are
skipped on the next run.
"""
import argparse
import json
import sqlite3
import sys
from pathlib import Path

import numpy as np

import config

FACES_SCHEMA = """
CREATE TABLE IF NOT EXISTS faces (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    file_id INTEGER NOT NULL,
    face_index INTEGER NOT NULL,
    embedding TEXT,
    bbox TEXT,
    best_similarity REAL,
    FOREIGN KEY(file_id) REFERENCES files(id)
);
CREATE INDEX IF NOT EXISTS idx_faces_file ON faces(file_id);
"""
# face_index = -1 is a marker row meaning "this photo was processed and had
# zero detected faces" - lets us skip it on incremental re-runs.


def imread_unicode(path):
    """cv2.imread breaks on some Windows paths (non-ASCII chars, long paths).
    Read bytes ourselves and decode, which is more robust."""
    import cv2
    try:
        data = np.fromfile(str(path), dtype=np.uint8)
        return cv2.imdecode(data, cv2.IMREAD_COLOR)
    except Exception:
        return None


def get_face_app():
    """Lazy-load insightface (slow import + first-run model download)."""
    from insightface.app import FaceAnalysis
    app = FaceAnalysis(name="buffalo_l", providers=["CPUExecutionProvider"])
    app.prepare(ctx_id=0, det_size=(config.FACE_DET_SIZE, config.FACE_DET_SIZE))
    return app


def get_embedding(face):
    """Prefer the pre-normalized embedding if the installed insightface
    version exposes it; otherwise normalize manually."""
    if hasattr(face, "normed_embedding") and face.normed_embedding is not None:
        return np.array(face.normed_embedding, dtype=np.float32)
    emb = np.array(face.embedding, dtype=np.float32)
    norm = np.linalg.norm(emb)
    return emb / norm if norm > 0 else emb


def cosine_sim(a, b):
    return float(np.dot(a, b))  # both already L2-normalized


def build_reference(app=None):
    ref_dir = Path(config.REFERENCE_PHOTOS_DIR)
    if not ref_dir.exists() or not any(ref_dir.iterdir()):
        print(f"[ERROR] No reference photos found in {ref_dir}/. "
              f"Add 5-10 clear photos of the person to find, then re-run.")
        sys.exit(1)

    app = app or get_face_app()
    embeddings = []
    photo_files = [p for p in ref_dir.iterdir()
                   if p.suffix.lower() in {".jpg", ".jpeg", ".png", ".bmp", ".webp"}]

    if not photo_files:
        print(f"[ERROR] No supported image files in {ref_dir}/ "
              f"(jpg/jpeg/png/bmp/webp).")
        sys.exit(1)

    for photo in photo_files:
        img = imread_unicode(photo)
        if img is None:
            print(f"[WARN] Could not read {photo}, skipping.")
            continue
        faces = app.get(img)
        if len(faces) == 0:
            print(f"[WARN] No face detected in {photo.name}, skipping.")
            continue
        if len(faces) > 1:
            print(f"[WARN] {photo.name} has {len(faces)} faces detected - "
                  f"using the largest one. For best results, use photos "
                  f"with only one person in them.")
            faces = sorted(faces, key=lambda f: (f.bbox[2]-f.bbox[0])*(f.bbox[3]-f.bbox[1]), reverse=True)
        embeddings.append(get_embedding(faces[0]).tolist())
        print(f"  [OK] {photo.name}")

    if not embeddings:
        print("[ERROR] Could not extract any usable face embeddings from "
              "the reference photos. Try clearer, more front-facing photos.")
        sys.exit(1)

    with open(config.REFERENCE_EMBEDDINGS_FILE, "w") as f:
        json.dump(embeddings, f)

    print(f"\nReference built from {len(embeddings)} photo(s), "
          f"saved to {config.REFERENCE_EMBEDDINGS_FILE}")


def load_reference():
    ref_file = Path(config.REFERENCE_EMBEDDINGS_FILE)
    if not ref_file.exists():
        print(f"[ERROR] {ref_file} not found. Run with --build-reference first.")
        sys.exit(1)
    with open(ref_file) as f:
        embeddings = json.load(f)
    return [np.array(e, dtype=np.float32) for e in embeddings]


def best_match_score(embedding, reference_embeddings):
    return max(cosine_sim(embedding, ref) for ref in reference_embeddings)


def scan_library():
    reference_embeddings = load_reference()
    conn = sqlite3.connect(config.DB_PATH)
    conn.executescript(FACES_SCHEMA)
    cur = conn.cursor()

    cur.execute("SELECT id, path FROM files WHERE category='Photo' AND error IS NULL")
    all_photos = cur.fetchall()

    cur.execute("SELECT DISTINCT file_id FROM faces")
    already_processed = {row[0] for row in cur.fetchall()}

    to_process = [(fid, path) for fid, path in all_photos if fid not in already_processed]
    print(f"Total photos in inventory : {len(all_photos)}")
    print(f"Already processed          : {len(already_processed)}")
    print(f"To process this run        : {len(to_process)}\n")

    if not to_process:
        print("Nothing new to scan.")
        conn.close()
        return

    app = get_face_app()
    processed, errors, total_faces = 0, 0, 0

    for file_id, path in to_process:
        img = imread_unicode(Path(path))
        if img is None:
            cur.execute(
                "INSERT INTO faces (file_id, face_index, embedding, bbox, best_similarity) VALUES (?,-1,NULL,NULL,NULL)",
                (file_id,),
            )
            errors += 1
            continue

        faces = app.get(img)
        if len(faces) == 0:
            cur.execute(
                "INSERT INTO faces (file_id, face_index, embedding, bbox, best_similarity) VALUES (?,-1,NULL,NULL,NULL)",
                (file_id,),
            )
        else:
            for i, face in enumerate(faces):
                emb = get_embedding(face)
                score = best_match_score(emb, reference_embeddings)
                bbox = json.dumps([float(x) for x in face.bbox])
                cur.execute(
                    "INSERT INTO faces (file_id, face_index, embedding, bbox, best_similarity) VALUES (?,?,?,?,?)",
                    (file_id, i, json.dumps(emb.tolist()), bbox, score),
                )
                total_faces += 1

        processed += 1
        if processed % 100 == 0:
            conn.commit()
            print(f"  ...{processed}/{len(to_process)} photos processed")

    conn.commit()
    conn.close()

    print(f"\n--- Scan complete ---")
    print(f"Photos processed : {processed}")
    print(f"Unreadable photos: {errors}")
    print(f"Total faces found: {total_faces}")


def rescan_similarity():
    """Recompute best_similarity for all previously-detected faces against
    the current reference embeddings, without re-running face detection.
    Useful after rebuilding the reference set with different photos."""
    reference_embeddings = load_reference()
    conn = sqlite3.connect(config.DB_PATH)
    cur = conn.cursor()
    cur.execute("SELECT id, embedding FROM faces WHERE face_index >= 0")
    rows = cur.fetchall()

    for face_id, emb_json in rows:
        emb = np.array(json.loads(emb_json), dtype=np.float32)
        score = best_match_score(emb, reference_embeddings)
        cur.execute("UPDATE faces SET best_similarity=? WHERE id=?", (score, face_id))

    conn.commit()
    conn.close()
    print(f"Recomputed similarity for {len(rows)} detected face(s).")


def export_matches(threshold):
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Font

    conn = sqlite3.connect(config.DB_PATH)
    df = pd.read_sql_query(
        """SELECT f.path, f.category, MAX(fc.best_similarity) as best_similarity
           FROM faces fc JOIN files f ON f.id = fc.file_id
           WHERE fc.face_index >= 0
           GROUP BY f.id
           HAVING best_similarity >= ?
           ORDER BY best_similarity DESC""",
        conn, params=(threshold,),
    )
    conn.close()

    out_dir = Path(config.OUTPUT_DIR)
    out_dir.mkdir(exist_ok=True)
    out_file = out_dir / "my_photos.xlsx"
    df.to_excel(out_file, index=False, sheet_name="Matches")

    wb = load_workbook(out_file)
    ws = wb["Matches"]
    for cell in ws[1]:
        cell.font = Font(name="Arial", bold=True)
    ws.freeze_panes = "A2"
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 70)
    wb.save(out_file)

    print(f"\nThreshold used      : {threshold}")
    print(f"Matching photos      : {len(df)}")
    print(f"Report exported to   : {out_file}")
    if len(df) > 0:
        print(f"Similarity range     : {df['best_similarity'].min():.3f} - {df['best_similarity'].max():.3f}")
    print("\nIf you're getting too many false positives, raise the threshold.")
    print("If real matches are missing, lower it. Re-check without rescanning:")
    print("    python find_faces.py --rescan-similarity --threshold <value>")


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--build-reference", action="store_true",
                         help="(Re)build reference embeddings from REFERENCE_PHOTOS_DIR")
    parser.add_argument("--rescan-similarity", action="store_true",
                         help="Recompute match scores for already-detected faces against current reference, without re-detecting")
    parser.add_argument("--threshold", type=float, default=None,
                         help="Override FACE_MATCH_THRESHOLD from config.py for this run's export")
    args = parser.parse_args()

    threshold = args.threshold if args.threshold is not None else config.FACE_MATCH_THRESHOLD

    if args.build_reference:
        build_reference()
    elif args.rescan_similarity:
        rescan_similarity()
        export_matches(threshold)
    else:
        scan_library()
        export_matches(threshold)
